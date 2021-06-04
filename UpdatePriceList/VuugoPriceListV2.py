import csv
import openpyxl
from time import sleep
VERSION = 2.5

# item class to hold item information

# item object that will hold all relevant information on each specified item


class item:
    def __init__(self, type1, type2, manu, quantity, price, name):
        self.type1 = type1
        self.type2 = type2
        self.manu = manu
        self.quantity = quantity
        self.price = price
        self.name = name

        # function to show the info in the requested instance
    def returnValues(self):
        return (self.type1, self.type2, self.manu, self.quantity, self.price, self.name)


def printDiv():
    print("-------------------------------------------------------")

    # used for splitting the item desc. name into relevant parts


def splitType(s):
    s1 = s[:s.find(":")]  										# gets the item's heading (type1)
    # gets the item's subheading (type2)
    s2 = s[(len(s1) + 1):(s.find(":", len(s1) + 1))]
    return s1, s2


# prints a loading bar curr = current # of items, total = total # of items, total_div = number of dividers
def loadingBar(curr, total, total_div):
    div = total//total_div  									# calculate how many items in each divider
    if curr < (div * total_div):  								# if the current number is less than 100%
        # check if a new divider is needed (everytime it hits div amount of items)
        if curr % div == 0:
            print("|{}{}| {:.1f}%".format(u"\u2588" *
                                          (curr//div), " " *
                                          (total_div - curr//div),
                                          curr/total * 100), end='\r')  						# prints the loading bar, then moves the cursor back to the beginning using carriage return
    # check if the current number is the maximum (100%) number
    elif curr == (div * total_div):
        print(" " * 100, end='\r')  							# clears the line
        print("|{}| 100.0%".format(u"\u2588" * total_div)
              )  												# prints a full loading bar with 100%
        # if it goes over the maximum (100%) amount, do nothing (this is possible because div * total_div may be less than total)
    sleep(0.005)  												# artifical pause

    return


# CSV File index constants
QB_NAME_COL = 29
QB_MANU_NAME_COL = 22
QB_QUANTITY_COL = 12
QB_PRICE1_COL = 25
QB_ITEM_TYPE_COL = 3

# print version
print("VERSION {}".format(VERSION))

# try to open "pricelist.csv"
opened = False
while not opened:
    try:
        csvFile = csv.reader(open("pricelist.CSV", 'r'))
        print("CSV file loaded successfully.")
        opened = True
    except:
        input("pricelist.csv does not exist. Put pricelist.csv in the same folder and press enter to try again.")

    # create, open, and name excel file
    wb = openpyxl.Workbook()
    pl = wb.active
    pl.title = "Vuugo Price List"

items = []  # list of item objects
errors = []  # list of errored items
# defines special types that must have subtypes (type2)
special_types = ["Processors", "Memory", "SSD", "Motherboard", "Video Card"]
# going through the csv file to find items that match requirements (>0 stock) and creates associated object
for row in csvFile:
    if len(row[QB_NAME_COL]) > 0:  # check if there exists a item list name for this row
        # check if the quantity is not 0 (neg is ok)
        if row[QB_QUANTITY_COL] != "0":
            # set the itemType (i use a diff variable because im splitting it later using splitType function)
            itemType = row[QB_ITEM_TYPE_COL]
            if ":" in itemType:  # checks to see if the item desc. is split into different parts
                type1, type2 = splitType(itemType)  # see splitType func
            else:  # if not, just use what is there and ignore type2
                type1 = itemType
                type2 = None
            manu = row[QB_MANU_NAME_COL]
            name = row[QB_NAME_COL]
            quantity = row[QB_QUANTITY_COL]
            price = row[QB_PRICE1_COL]
            # if there is a / in the price (e.g. 99/0504) then remove whats after / (inclusive)
            if "/" in price:
                price = price[:price.find("/")]
            # append item object to items list
            items.append(item(type1, type2, manu, quantity, price, name))

# deletes first item object (useless as it contains the titles only "('Item', 'None', 'Manu. Part No.', 'Quantity On Hand', 'Price L', 'itemname')")
del items[0]

# set current type1 (in this case it is the very first one)
type1 = items[0].type1

# set current type2 (in this case it is the very first one)
type2 = items[0].type2

x, y = 1, 1  # creates xy coordinate that will act as a 'cursor'

titles = ["Manufacturer's Number", "Quantity",
          "Price", "Item Name"]  # title names
item_num = 0
# loop through the title names and place them beside eachother on y = 1
for title_index in range(len(titles)):
    pl.cell(row=y, column=title_index + 1,
            value=titles[title_index]).font = openpyxl.styles.Font(size=14)
y += 1  # moves the cursor down one line
pl.cell(row=y, column=x, value=type1).font = openpyxl.styles.Font(
    size=14, bold=True)  # place the first type1 title on the second line
y += 1

for i in items:  # loops through each item object and assigns their position
    item_num += 1
    try:
        # create a list of item information (to loop through later)
        itemInfo = [i.manu, int(i.quantity), float(i.price), i.name]
    except ValueError as v:
        # clears the current line (may have leftovers from loadingBar function)
        print(" " * (100), end='\r')
        print("(!!!) {} HAS NO PRICE".format(i.name))
        itemInfo = [i.manu, int(i.quantity), "Call", i.name]
        errors.append(i)
        i.error = v
    if i.type1 != type1:  # since the type1s will be in order, once it changes the cursor will move down 2, place the item type1 title, and then move one down
        y += 2  # move 2 lines down
        type1 = i.type1  # set new current type
        # place the type1 title as well as make it bold and higher font size
        pl.cell(row=y, column=x, value=type1).font = openpyxl.styles.Font(
            size=14, bold=True)
        y += 1
    # if the current item's type requries a subheading (type2)
    if i.type1 in special_types:
        if type2 != i.type2:  # they are in order, so only create a new subheading if the type2 changes
            type2 = i.type2  # set new current type2
            y += 1
            pl.cell(row=y, column=x, value=type2).font = openpyxl.styles.Font(
                bold=True)  # place new type2 subheading and bold it
            y += 1
    # loops through the object's information (list above)
    for info in range(len(itemInfo)):
        pl.cell(row=y, column=info + 1, value=itemInfo[info])
    loadingBar(item_num, len(items), 25)
    y += 1
print("Loaded {:d} items".format(item_num))


printDiv()
if len(errors) > 0:  # check if there are any errored items
    print("The following items do not have a price (currently has 'call'): ")
    for i in errors:
        print("{:s} ({:s})".format(i.name, i.type1))
    printDiv()

saved = False
show_error_list = ["showerrors", "displayerrors", "showerror", "show errors", "display errors",
                   "display error", "outputerror", "outputerrors", "output error", "output errors"]
while not saved:  # saves the file
    saveName = input("save vuugo price list as: ")
    if saveName.lower() in show_error_list:
        for i in errors:
            print("{} -> [{}]".format(i.name, i.error))
    else:
        saveName += '.xlsx'
        try:
            wb.save(saveName)
            print("saved successfully as {}".format(saveName))
            saved = True
        except:
            print("(!!!) could not save! (!!!)")
input("press enter to exit")
