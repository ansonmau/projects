import csv
import openpyxl
 

# to be seperated: cpu (i 10) , memory (i 6), ssd (i 12), mb (i 8), videocard (i 13)

VERSION = 1.3

QB_NAME_COL = 29
QB_MANU_NAME = 22
QB_QUANTITY_COL = 12
QB_PRICE1_COL = 25
QB_ITEM_TYPE_COL = 3
titles = ["Manufacturer's Number", "Quantity", "Price", "Item Name"]
item_types = ["Barebone", "UPS", "Accessories", "Fans/cooler", "Flash Drive/Memory", "Hard Drive", "Memory", "Monitor", "Motherboard", "Power Supply", "Processors", "Software", "SSD", "Video Card"]
items = [[[]] for _ in item_types]

print("VERSION={}".format(VERSION))
while True:
	try:
		csvFile = csv.reader(open("pricelist.CSV", 'r'))
		print("[pricelist.csv] opened successfully")
		break
	except:
		print("[pricelist.csv] does not exist in current directory")
		input("press enter to try again...")

wb = openpyxl.Workbook()
pricelist = wb.active
pricelist.title = "Price List"
csv_total = 0
csv_found = 0
csv_added = 0
csvFile = list(csvFile)[1:]
index = 0

for row in csvFile: # iterate through each row of csv file
	csv_total += 1
	if len(row[QB_NAME_COL]) > 1:
		csv_found += 1
		if row[QB_QUANTITY_COL] != '0':
			itemTypeOG = row[QB_ITEM_TYPE_COL]
			if ":" in itemType:
				itemType = itemTypeOG[:itemTypeOG.index(":")]
			if itemType in item_types:
				name = row[QB_NAME_COL]
				manu_name = row[QB_MANU_NAME].strip()
				quantity = int(row[QB_QUANTITY_COL])
				price = row[QB_PRICE1_COL]
				if '/' in price:
					price = price[:price.index("/")]
				price = float(price)
				items[item_types.index(itemType)].append([manu_name, quantity, price, name])
				csv_added += 1
				print("added {} under {}".format(name, itemType))
			else:
				print("{} not a proper item type".format(itemType))


for i in range(len(titles)):
	pricelist.cell(row=1, column=i+1, value=titles[i])

x=1
y=3
typeIndex = 0
for itemList in items:
	pricelist.cell(row=y, column=x, value=item_types[typeIndex])
	pricelist.cell(row=y, column=x).font = openpyxl.styles.Font(bold=True)
	y+=1
	for item in itemList:
		for spec in item:
			pricelist.cell(row=y, column=x, value=spec)
			x+=1
		x=1
		y+=1
	y+=2
	typeIndex+=1
print("[ {} / {} / {} ]".format(csv_added, csv_found, csv_total))

saved = False
while not saved:
	saveName = input("save vuugo price list as: ")
	saveName += '.xlsx'
	try:
		wb.save(saveName)
		print("saved successfully as {}".format(saveName))
		saved = True
	except:
		print("(!!!) could not save! (!!!)")
input("press enter to exit")